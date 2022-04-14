# This project we will update an Excel file with automations

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# This makes calling on the file shorter ^^^ and import a chart feature

# We are calling on the worksheet and storing it in a variable vvv
wb = xl.load_workbook('transactions.xlsx')

# We are using the first Sheet in our xl File. The capital letter is important
sheet = wb['Sheet1']

# We can work directly with cells and row in our file.
cell = sheet['a1']

# Can also write this as "cell = sheet.cell(1, 1)" ^^^

# This will show us how many rows we have in our excel
"print(sheet.max_row)"

# working with our sheet, we want to work in certain rows.
# This will have us working from row 2 to row 5
# We want to fix/update our current prices on our sheet.
# lastly we are adding a new row to store our new corrected_prices
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# This will call on the imported chart file above to create a graph
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

# We stored the variable values with the specific area we want to see on the graph
# we called on the BarChart function and assigned it a variable
# we added the chart and told it to only use what's in 'values
# lastly we told xl where we wanted the BarChart


# This will add a new saved file with all the updates above
wb.save('transactions2.xlsx')
